#! /usr/bin/python3

from datetime import date
import os
import sys
import chevron
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

def extract_test_case(sheet_ranges : Worksheet):
    test_case = {}
    test_case['id'] = sheet_ranges['C2'].value
    if test_case['id'] is None:
        return None
    test_case['name'] = sheet_ranges['C3'].value
    extract_generic_data(sheet_ranges, test_case, start_row=4)
    return test_case

def extract_test_specification(sheet_ranges):
    test_specification = {}
    test_specification['id'] = sheet_ranges['C2'].value
    if test_specification['id'] is None:
        return None
    test_specification['parent_reference'] = sheet_ranges['C3'].value
    test_specification['name'] = sheet_ranges['C4'].value
    extract_generic_data(sheet_ranges, test_specification, start_row=6, diagram_root_path='..')
    return test_specification


def extract_experiment_specification(sheet_ranges):
    exp_spec = {}
    exp_spec['id'] = sheet_ranges['C2'].value
    if exp_spec['id'] is None:
        return None
    exp_spec['parent_reference'] = sheet_ranges['C3'].value
    exp_spec['name'] = sheet_ranges['C4'].value
    extract_generic_data(sheet_ranges, exp_spec, start_row=6, diagram_root_path=os.path.join('..', '..'))
    return exp_spec


def extract_generic_data(sheet_ranges : Worksheet, object, start_row=1, end_row=None, diagram_root_path='.'):
    if not 'subsections' in object:
        object['subsections'] = []

    for row_idx in range(start_row, 1000):
        if str(sheet_ranges['A' + str(row_idx)].value).lower() == 'diagrams':
            break
        
        headline_cell = sheet_ranges['B' + str(row_idx)]
        if headline_cell.font.bold:
            section = {
                'section_title': headline_cell.value,
                'subsections': []
            }
            object['subsections'].append(section)
            if is_gray(get_cell_right(headline_cell)):
                continue
            else:
                section['contents'] = get_cell_right(headline_cell).value
        elif str(headline_cell.value).lower() == 'description':
            section['contents'] = get_cell_right(headline_cell).value
        elif str(headline_cell.value).lower() == 'diagram reference':
            section['diagrams'] = extract_diagrams(sheet_ranges, get_cell_right(headline_cell).value, diagram_root_path=diagram_root_path)
        elif headline_cell.value:
            subsection = {
                'section_title': headline_cell.value,
                'contents': get_cell_right(headline_cell).value
            }
            section['subsections'].append(subsection)

    return object

def extract_diagrams(sheet : Worksheet, diagram_reference, diagram_root_path='.'):
    return_list = []

    if not diagram_reference:
        return return_list

    diagram_references = [x.strip() for x in diagram_reference.split(';')]

    for i in range(1, 1000):
        if str(sheet['A' + str(i)].value).lower().strip() == 'diagrams':
            diagram_id_row = i + 1

    for dia_ref in diagram_references:
        col = 3
        while sheet.cell(row=diagram_id_row, column=col).value is not None:
            if sheet.cell(row=diagram_id_row, column=col).value == dia_ref:
                return_list.append(
                    {
                        'diagram_name': sheet.cell(row=diagram_id_row+1, column=col).value,
                        'diagram_uri': os.path.join(diagram_root_path, sheet.cell(row=diagram_id_row+3, column=col).value)
                    })
                break
            col += 1

    return return_list

def get_cell_below(cell : Cell) -> Cell:
    ws = cell.parent
    return ws[cell.column_letter() + str(cell.row + 1)]

def get_cell_right(cell : Cell) -> Cell:
    ws = cell.parent
    return ws.cell(cell.row, cell.column + 1)

def is_gray(cell : Cell):
    return type(cell.fill.fgColor.theme) == int

def extract_table(start_cell : Cell, object):
    table_columns = []
    title_cell = start_cell
    while not title_cell.font.bold:
        column = []
        cell = title_cell
        while cell.value:
            column.append(cell.value)
            cell = get_cell_right(cell)            
        table_columns.append(column)

        title_cell = get_cell_below(title_cell)

    if len(table_columns) > 0:
        object['table'] = table_columns

def escape_quotes(text):
    text = str(text)
    text = text.replace('\'', '\'\'')
    return text

def url_safe(text):
    text = str(text)
    text = "".join([c for c in text if c.isalpha() or c.isdigit()]).rstrip()
    return text
        
def add_header(object, title, link_title, date, description):
    object['title'] = escape_quotes(title)
    object['linkTitle'] = escape_quotes(link_title)
    object['date'] = escape_quotes(date)
    object['description'] = escape_quotes(description)
    return object

def main(filename, output_dir):
    try:
        wb = load_workbook(filename)
    except:
        print("File does not exist!")
        return

    test_case = None
    test_specifications = []
    experiment_specifications = []
    for sheet in wb:
        sheet_ranges = wb[sheet.title]
        sheet_type = sheet_ranges['A1']
        if sheet_type.value == 'Test Case':
            test_case = extract_test_case(sheet_ranges)
        elif sheet_type.value == 'Test Specification':
            test_specifications.append(extract_test_specification(sheet_ranges))
        elif sheet_type.value == 'Experiment Specification':
            experiment_specifications.append(extract_experiment_specification(sheet_ranges))

    test_specifications = [ts for ts in test_specifications if ts is not None]
    experiment_specifications = [es for es in experiment_specifications if es is not None]

    output_files = {}

    if test_case:
        mtime = date.fromtimestamp(os.path.getmtime(filename)).isoformat()
        test_case_filename = os.path.splitext(os.path.basename(filename))[0]
        add_header(test_case, 'Test Case ' + test_case_filename, test_case_filename, mtime, test_case['name'])
        with open('TestCase.mustache', 'r') as template:
            md_test_case = chevron.render(template=template, data=test_case)
            index_name = 'index.md' if len(test_specifications) == 0 else '_index.md'
            output_files['root'] = {
                'dir_path': os.path.join('.'), 
                'filename': index_name,
                'content': md_test_case
            }
            # print(md_test_case)
    
        for ts in test_specifications:
            add_header(ts, 'Test Specification ' + ts['id'], ts['id'], mtime, ts['name'])
            with open('TestSpecification.mustache', 'r') as template:
                md_test_spec = chevron.render(template=template, data=ts)
                output_files[ts['id']] = {
                    'dir_path': os.path.join('.', ts['id']),
                    'filename': 'index.md', 
                    'content': md_test_spec
                }

        for es in experiment_specifications:
            add_header(es, 'Experiment Specification ' + es['id'], es['id'], mtime, es['name'])
            parent_path = None
            if es['parent_reference'] in output_files:
                parent_path = es['parent_reference']
                output_files[es['parent_reference']]['filename'] = '_index.md'

            if parent_path:
                with open('ExperimentSpecification.mustache', 'r') as template:
                    md_exp_spec = chevron.render(template=template, data=es)
                    output_files[es['id']] = {
                        'dir_path': os.path.join('.', parent_path, es['id']),
                        'filename': 'index.md', 
                        'content': md_exp_spec
                    }

    for of in output_files.values():
        file_path = os.path.join(output_dir, of['dir_path'], of['filename'])
        print('Creating file ' + file_path)
        if not os.path.exists(os.path.dirname(file_path)):
            os.makedirs(os.path.dirname(file_path))
        
        with open(file_path, 'w', encoding='utf-8') as fs:
            fs.write(of['content'])
        

#Python3
if __name__ == '__main__':
    output_dir = '.'
    if len(sys.argv) > 1:
        filename = str(sys.argv[1])
        if len(sys.argv) > 2:
            output_dir = str(sys.argv[2])
    else:
        print("No arguments introduced")

    main(filename, output_dir)
